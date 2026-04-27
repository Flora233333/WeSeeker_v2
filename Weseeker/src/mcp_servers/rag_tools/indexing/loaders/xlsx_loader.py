from __future__ import annotations

from pathlib import Path

from langchain_core.documents import Document
from openpyxl import load_workbook

_MAX_ROWS_PER_SHEET = 50


def load_xlsx(path: Path) -> list[Document]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    documents: list[Document] = []

    try:
        for worksheet in workbook.worksheets:
            rows: list[list[str]] = []
            for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if row_index > _MAX_ROWS_PER_SHEET:
                    break
                values = ["" if cell is None else str(cell).strip() for cell in row]
                if any(values):
                    rows.append(values)

            if not rows:
                continue

            table_text = _rows_to_markdown(rows)
            documents.append(
                Document(
                    page_content=table_text,
                    metadata={
                        "doc_type": "xlsx",
                        "file_path": path.as_posix(),
                        "file_name": path.name,
                        "file_ext": path.suffix.lower(),
                        "sheet_name": worksheet.title,
                    },
                )
            )
    finally:
        workbook.close()

    return documents


def _rows_to_markdown(rows: list[list[str]]) -> str:
    header = rows[0]
    body = rows[1:]
    header_line = "| " + " | ".join(header) + " |"
    separator_line = "| " + " | ".join(["---"] * len(header)) + " |"
    body_lines = ["| " + " | ".join(row) + " |" for row in body]
    return "\n".join([header_line, separator_line, *body_lines])
